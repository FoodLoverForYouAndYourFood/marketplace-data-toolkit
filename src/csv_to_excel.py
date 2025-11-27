import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


def convert_csv_to_xlsx(csv_path: Path, xlsx_path: Path) -> None:
    """Convert a UTF-8 CSV file to XLSX using openpyxl."""
    csv_path = Path(csv_path)
    xlsx_path = Path(xlsx_path)
    xlsx_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active

    with csv_path.open(encoding="utf-8") as fh:
        reader = csv.reader(fh)
        for row in reader:
            ws.append(row)

    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(length + 2, 10), 60)

    wb.save(xlsx_path)


__all__ = ["convert_csv_to_xlsx"]
