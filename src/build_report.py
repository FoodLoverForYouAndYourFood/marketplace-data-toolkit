import argparse
import csv
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADERS = [
    "ID продавца",
    "Продавец",
    "Предмет",
    "НМ",
    "Бренд",
    "Наименование",
    "Ссылка WB",
    "Ссылка OZON",
    "Цена WB 20.10",
    "Цена OZON 20.10",
    "Маржа 20.10",
    "Цена WB 10.11",
    "Цена OZON 10.11",
    "Дельта",
    "КАМ",
]


def load_csv(path: Optional[Path]) -> Dict[str, Dict[str, str]]:
    if not path or not path.exists():
        return {}
    with path.open(encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        return {row["product_id"]: row for row in reader if row.get("product_id")}


def merge_products(
    wb_rows: Dict[str, Dict[str, str]],
    ozon_rows: Dict[str, Dict[str, str]],
) -> List[Dict[str, Optional[str]]]:
    keys = sorted(set(wb_rows.keys()) | set(ozon_rows.keys()))
    merged: List[Dict[str, Optional[str]]] = []
    for key in keys:
        wb_row = wb_rows.get(key, {})
        oz_row = ozon_rows.get(key, {})
        merged.append({"id": key, "wb": wb_row, "ozon": oz_row})
    return merged


def autosize_columns(ws) -> None:
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
        ws.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(length + 2, 12), 60)


def build_workbook(rows: List[Dict[str, Optional[str]]]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Лист1"
    ws.append(HEADERS)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for idx, row in enumerate(rows, start=2):
        wb_data = row["wb"] or {}
        oz_data = row["ozon"] or {}
        ws.append(
            [
                wb_data.get("supplier_id") or oz_data.get("supplier_id") or "",
                wb_data.get("supplier_name") or oz_data.get("supplier_name") or "",
                wb_data.get("subject_id") or oz_data.get("subject_id") or "",
                wb_data.get("product_id") or oz_data.get("product_id"),
                wb_data.get("brand") or oz_data.get("brand"),
                wb_data.get("name") or oz_data.get("name"),
                wb_data.get("url"),
                oz_data.get("url"),
                safe_float(wb_data.get("price")),
                safe_float(oz_data.get("price")),
                None,
                safe_float(wb_data.get("price")),
                safe_float(oz_data.get("price")),
                None,
                build_comment(wb_data, oz_data),
            ]
        )
        ws.cell(row=idx, column=11).value = f'=IF(OR(J{idx}=0,J{idx}=""),"",I{idx}/J{idx}-1)'
        ws.cell(row=idx, column=14).value = f'=IF(OR(M{idx}=0,M{idx}=""),"",L{idx}/M{idx}-1)'
        for col in range(1, len(HEADERS) + 1):
            ws.cell(row=idx, column=col).alignment = Alignment(vertical="top", wrap_text=True)

    ws.freeze_panes = "A2"
    autosize_columns(ws)
    return wb


def build_comment(wb_row: Dict[str, str], oz_row: Dict[str, str]) -> str:
    parts = []
    if wb_row.get("rating_value"):
        parts.append(f"WB рейтинг {wb_row['rating_value']} ({wb_row.get('review_count')} отзывов)")
    if oz_row.get("rating_value"):
        parts.append(f"OZ рейтинг {oz_row['rating_value']} ({oz_row.get('review_count')} отзывов)")
    return "; ".join(part for part in parts if part)


def safe_float(value: Optional[str]) -> Optional[float]:
    if not value:
        return None
    try:
        return float(str(value).replace(",", "."))
    except ValueError:
        return None


def main() -> None:
    parser = argparse.ArgumentParser(description="Построение Excel отчета в формате Примеры.xlsx")
    parser.add_argument("--wb-csv", type=Path, required=True, help="CSV с данными Wildberries.")
    parser.add_argument("--ozon-csv", type=Path, help="CSV с данными Ozon.")
    parser.add_argument("--out", type=Path, default=Path("report.xlsx"), help="Путь к результирующему Excel.")
    args = parser.parse_args()

    wb_rows = load_csv(args.wb_csv)
    oz_rows = load_csv(args.ozon_csv)
    rows = merge_products(wb_rows, oz_rows)
    workbook = build_workbook(rows)
    workbook.save(args.out)
    print(f"Report saved to {args.out}")


if __name__ == "__main__":
    main()
